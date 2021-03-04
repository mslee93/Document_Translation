import openpyxl
from openpyxl.utils.cell import get_column_letter
from settings import engine_dic
from _regex import check_word, check_whitespace


def translate(value, src_lang, tgt_lang, engine):
    if value is None:
        return False
    if type(value) == str:
        value = str(value)
        if value[0] == "=":
            return False
        if check_word.fullmatch(value):
            return False
        else:
            if check_whitespace.match(value):
                front_whitespace = check_whitespace.match(value)[0]
            else:
                front_whitespace = ''
            new_text = engine.translate(value, src_lang, tgt_lang)
            # print(value)
            if new_text is False:
                return False
            else:
                new_text = front_whitespace + new_text
                return new_text


# Change Sheet name
def translateSheetname(sheet, sheet_name, src_lang, tgt_lang, engine):
    new_sheet_name = translate(sheet_name, src_lang, tgt_lang, engine)
    if new_sheet_name:
        try:
            sheet.title = new_sheet_name[:31]
        except:
            pass


def translateCell(sheet, src_lang, tgt_lang, engine):
    amountOfRows = sheet.max_row
    amountOfColumns = sheet.max_column

    for i in range(amountOfColumns):
        for k in range(amountOfRows):
            if hasattr(sheet[get_column_letter(i+1)+str(k+1)], "value"):
                cell = sheet[get_column_letter(i+1)+str(k+1)].value
                new_cell = translate(cell, src_lang, tgt_lang, engine)
                if new_cell:
                    try:
                        sheet[get_column_letter(i + 1) + str(k + 1)] = new_cell
                    except:
                        pass


def run(src_file, tgt_file, src_lang, tgt_lang, eng):

    try:
        engine = engine_dic[eng]
    except:
        engine = None
    
    if engine is not None:
        wb = openpyxl.load_workbook(src_file)
        sheet_list = wb.sheetnames
        for sheet_name in sheet_list:
            sheet = wb[sheet_name]
            # Change Sheet name
            translateSheetname(sheet, sheet_name, src_lang, tgt_lang, engine)
            # Change Cells
            translateCell(sheet, src_lang, tgt_lang, engine)

        wb.save(tgt_file)

